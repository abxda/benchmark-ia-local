import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook

# Leer el archivo CSV
df = pd.read_csv('ventas.csv')

# Verificar que los datos se leyeron correctamente
print("Datos leídos del CSV:")
print(df)
print(f"\nColumnas: {list(df.columns)}")
print(f"Filas: {len(df)}")

# Calcular la suma de ventas por región
ventas_por_region = df.groupby('region')['ventas'].sum().reset_index()
ventas_por_region.columns = ['region', 'ventas_totales']

print(f"\nVentas totales por región:")
print(ventas_por_region)

# Crear el archivo Excel con una hoja llamada 'Resumen'
with pd.ExcelWriter('reporte.xlsx', engine='openpyxl') as writer:
    ventas_por_region.to_excel(writer, sheet_name='Resumen', index=False)

print(f"\nArchivo 'reporte.xlsx' creado correctamente.")

# Verificar el contenido del archivo Excel
wb = load_workbook('reporte.xlsx')
print(f"Hojas en el archivo: {wb.sheetnames}")
ws = wb['Resumen']
print(f"Celulas de la fila 1: {ws[1]}")
print(f"Celulas de la fila 2: {ws[2]}")
print(f"Celulas de la fila 3: {ws[3]}")

wb.close()

print("\nScript completado exitosamente.")
