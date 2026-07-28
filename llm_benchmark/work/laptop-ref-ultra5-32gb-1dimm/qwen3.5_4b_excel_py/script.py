import pandas as pd
from openpyxl import Workbook

# Leer datos desde CSV en directorio actual
df = pd.read_csv('ventas.csv')

# Calcular suma de ventas por región usando groupby y agg
resumen_regiones = df.groupby('region')['ventas'].sum().reset_index()

# Renombrar columnas para coincidir con el requisito (region, ventas_totales)
resumen_regiones.columns = ['region', 'ventas_totales']

# Crear nuevo libro de Excel
wb = Workbook()

# Seleccionar la primera hoja y renombrarla a "Resumen"
ws = wb.active
ws.title = "Resumen"

# Escribir encabezados en las celdas A1 y B1
ws.cell(row=1, column=1).value = 'region'
ws.cell(row=1, column=2).value = 'ventas_totales'

# Escribir datos de la columna 'region' (columna 0 del dataframe)
for idx, row in resumen_regiones.iterrows():
    ws.cell(row=idx + 2, column=1).value = row['region']

# Escribir datos de la columna 'ventas_totales' (columna 1 del dataframe)
for idx, row in resumen_regiones.iterrows():
    ws.cell(row=idx + 2, column=2).value = row['ventas_totales']

# Guardar el archivo Excel en directorio actual con nombre reporte.xlsx
wb.save('reporte.xlsx')
